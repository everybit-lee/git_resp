# -*- coding: utf-8 -*-

#import pymysql
#import xlwt
from sklearn import feature_extraction
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.feature_extraction.text import CountVectorizer
from openpyxl import Workbook

wb=Workbook()
ws=wb.active
#book = xlwt.Workbook(encoding='utf-8', style_compression=0)
#f01=open('./data/tfidf_res_jtext.csv',"w+")
f02=open("./data/jtext.csv","r")
#sheet = book.add_sheet('total', cell_overwrite_ok=True)
rs=f02.readlines()
print "len(rs)="+ str(len(rs))+"--------------"
vectorizer = CountVectorizer()  # 该类会将文本中的词语转换为词频矩阵，矩阵元素a[i][j] 表示j词在i类文本下的词频
transformer = TfidfTransformer()  # 该类会统计每个词语的tf-idf权值
tfidf = transformer.fit_transform(
    vectorizer.fit_transform(rs))  # 第一个fit_transform是计算tf-idf，第二个fit_transform是将文本转为词频矩阵
word = vectorizer.get_feature_names()  # 获取词袋模型中的所有词语
weight = tfidf.toarray()  # 将tf-idf矩阵抽取出来，元素a[i][j]表示j词在i类文本中的tf-idf权重
print "len(word)=" +str(len(word)) +"-----  and weight.shape="
print weight.shape
p=1
dict={}
k=1
for i in range(len(weight)):  # 打印每类文本的tf-idf词语权重，第一个for遍历所有文本，第二个for便利某一类文本下的词语权重
	#st="J"+str(i)
	#sheet.write(0,p,st)
	#sheet.write(0,p,'tfidf')
	for j in range(len(word)):
		if weight[i][j]>0:
			ws.cell(row=k,column=p).value=word[j]
			ws.cell(row=k, column=p+1).value = weight[i][j]
			#sheet.write(k, p, word[j])
			#sheet.write(k, p+1, weight[i][j])
			k = k + 1
			#dict[word[j]]=weight[i][j] #将结果放在一个字典中
			#print word[j],weight[i][j]
#print "len(dict)=  "  + str(len(dict))
#sortres = sorted(dict.items(), key=lambda item: item[1], reverse=True) # sort the dict
#print "len(sortres)=  " + str(len(sortres))
#
#for key in sortres:
	#print key[0]+","+key[1]
	#f01.write(key[0]+","+key[1])
	#f01.write("\n")
#	sheet.write(k, p, key[0])
#	sheet.write(k, p+1, key[1])
#
	#p=p+3
wb.save(filename="./data/tfidf-all-1130-3.xlsx")
# 最后，将以上操作保存到指定的Excel文件中
#book.save("./data/tfidf-all-1130-3.xlsx")
#book.save(r"E:\edu_data\datatfidf-total.xls")
f02.close()
#f01.close()
print "done!--------------------"